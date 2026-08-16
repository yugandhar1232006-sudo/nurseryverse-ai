"""
Format renderers for `ReportGenerationService` -- one function per
`ReportFormat`, each taking the exact same shape (a title, a flat list of
column headers, and a list of already-serialized row tuples) so every one
of this module's ~17 report types renders through the same four functions
instead of each report type growing its own bespoke PDF/Excel/CSV/JSON
code (the "No duplicated reporting logic" QUALITY requirement, applied to
the export layer, not just the query layer).

`reportlab`/`openpyxl` have been listed in requirements/base.txt since
Phase 5 but never imported by any module before this one -- Module 12 is
this codebase's first actual PDF/Excel generation code.
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any


def _json_safe(value: Any) -> Any:
    """
    Shared scalar coercion: enums to `.value`, Decimals to float, dates to
    ISO strings, UUIDs to strings. The UUID case matters specifically for
    `render_excel` -- every report type's `rows` here carries raw
    `uuid.UUID` id columns straight from the ORM, and unlike `csv.writer`
    (calls `str()` on anything) or `json.dumps` (has this module's own
    `default=str` fallback), `openpyxl` raises `ValueError` on any value
    it doesn't recognize as a primitive -- caught by this module's own
    export smoke test before it could reach a real report generation run.
    """
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    return value


def render_csv(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_json_safe(v) for v in row])
    return buffer.getvalue().encode("utf-8")


def render_json(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    records = [dict(zip(headers, [_json_safe(v) for v in row])) for row in rows]
    payload = {"title": title, "generated_at": datetime.utcnow().isoformat(), "row_count": len(records), "rows": records}
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def render_excel(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    workbook = Workbook()
    # Phase 6 Module 14 (Production Readiness) defect fix: `Workbook.active`
    # is typed `Worksheet | Chartsheet | None` (types-openpyxl, installed
    # this module to close a `mypy app` gap -- see requirements/dev.txt) --
    # a real over-approximation for this call site specifically, since a
    # freshly constructed `Workbook()` always creates and activates one
    # default plain Worksheet (openpyxl's own documented behavior; a
    # Chartsheet or a missing active sheet only happens after explicit
    # sheet manipulation this function never does). The `isinstance` check
    # below is a genuine runtime guard (not just a type-checker hint) that
    # also lets mypy narrow `sheet` to `Worksheet` for every subsequent
    # line in this function, none of which are valid on a Chartsheet.
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise RuntimeError("openpyxl Workbook() did not create the expected default Worksheet")
    sheet.title = title[:31] or "Report"  # Excel sheet names are capped at 31 characters

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([_json_safe(v) for v in row])

    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        # Phase 6 Module 14 (Production Readiness) defect fix: was
        # `column_cells[0].column_letter` -- `.column_letter` genuinely
        # does not exist on `MergedCell` (confirmed against the real
        # openpyxl class, not just its stubs: only plain `Cell` has it,
        # `MergedCell` only exposes the numeric `.column`). This sheet
        # never actually merges any cells, so `column_cells[0]` is always
        # a plain `Cell` in practice, but `sheet.columns`'s general type
        # is `Cell | MergedCell` and mypy correctly can't know that from
        # here -- `get_column_letter(int)` works identically for both
        # (`.column` is defined on both), so this fix is both type-correct
        # and has zero behavior change from the original.
        column_index = column_cells[0].column
        if column_index is None:
            continue  # not expected for an appended, non-merged cell -- skip rather than guess a width
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = min(max(length + 2, 10), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_pdf(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    # `landscape` -- most report tables here run wide (10+ columns for
    # things like Plant Reports/Inventory Reports); portrait would just
    # wrap every table into an unreadable mess.
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title=title)
    styles = getSampleStyleSheet()

    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    elements.append(
        Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} -- {len(rows)} rows", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    table_data = [headers] + [[str(_json_safe(v)) if v is not None else "" for v in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F8E9")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


_RENDERERS = {
    "PDF": (render_pdf, "pdf", "application/pdf"),
    "EXCEL": (render_excel, "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "CSV": (render_csv, "csv", "text/csv"),
    "JSON": (render_json, "json", "application/json"),
}


def render(*, format_name: str, title: str, headers: list[str], rows: list[list[Any]]) -> tuple[bytes, str, str]:
    """Returns (file bytes, file extension, content type) for `format_name` (a `ReportFormat` member name)."""
    entry = _RENDERERS.get(format_name)
    if entry is None:
        raise ValueError(f"Unsupported report format: {format_name}")
    renderer, extension, content_type = entry
    content = renderer(title=title, headers=headers, rows=rows)
    return content, extension, content_type
