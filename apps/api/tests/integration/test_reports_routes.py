"""
Integration tests for Module 12's REST API (app/api/routes/reports.py):
authentication, authorization (`reports:read` vs `reports:export`),
validation failures, pagination, filtering, tenant isolation, branch
isolation, report generation (background task + status polling),
download, failed reports, scheduled-report CRUD/pause/resume/run-due, and
all four export formats -- end to end through the real ASGI app, the same
`auth_client`/`authenticated_client` harness-backed split every other
module's route tests use (see tests/integration/test_notification_routes.py).
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timedelta, timezone

import pytest

from app.db.enums import ReportFormat, ReportScheduleFrequency, ReportStatus, ReportType

pytestmark = pytest.mark.integration


def _grant_read(harness, user, *, org_id, branch_ids=None):
    harness.grant_role(user, org_id=org_id, role_code="viewer", permission_codes=["reports:read"], branch_ids=branch_ids)


def _grant_export(harness, user, *, org_id, branch_ids=None):
    harness.grant_role(
        user, org_id=org_id, role_code="owner", permission_codes=["reports:read", "reports:export"], branch_ids=branch_ids
    )


def _future(days: int = 1) -> str:
    return (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()


# --------------------------------------------------------------------------
# Authentication -- every route requires a bearer token
# --------------------------------------------------------------------------

_PLACEHOLDER_ID = "00000000-0000-0000-0000-000000000000"

_UNAUTH_ROUTES = [
    ("get", "/api/v1/dashboards/executive"),
    ("get", "/api/v1/dashboards/nursery"),
    ("get", f"/api/v1/dashboards/branch/{_PLACEHOLDER_ID}"),
    ("get", "/api/v1/dashboards/plant"),
    ("get", "/api/v1/dashboards/inventory"),
    ("get", "/api/v1/dashboards/sales"),
    ("get", "/api/v1/dashboards/customer"),
    ("get", "/api/v1/dashboards/ai"),
    ("get", "/api/v1/dashboards/financial"),
    ("get", "/api/v1/analytics/kpi-summary"),
    ("get", "/api/v1/analytics/revenue-trend"),
    ("get", "/api/v1/analytics/growth-trend"),
    ("get", "/api/v1/analytics/inventory-trend"),
    ("get", "/api/v1/analytics/plant-health-trend"),
    ("get", "/api/v1/analytics/sales-forecast"),
    ("get", "/api/v1/analytics/disease-trend"),
    ("get", "/api/v1/analytics/customer-analytics"),
    ("get", "/api/v1/analytics/employee-productivity"),
    ("get", "/api/v1/analytics/branch-performance"),
    ("get", "/api/v1/reports/catalog"),
    ("get", "/api/v1/reports/scheduled"),
    ("post", "/api/v1/reports/scheduled"),
    ("post", "/api/v1/reports/scheduled/run-due"),
    ("get", f"/api/v1/reports/scheduled/{_PLACEHOLDER_ID}"),
    ("patch", f"/api/v1/reports/scheduled/{_PLACEHOLDER_ID}"),
    ("post", f"/api/v1/reports/scheduled/{_PLACEHOLDER_ID}/pause"),
    ("post", f"/api/v1/reports/scheduled/{_PLACEHOLDER_ID}/resume"),
    ("delete", f"/api/v1/reports/scheduled/{_PLACEHOLDER_ID}"),
    ("post", "/api/v1/reports"),
    ("get", f"/api/v1/reports/{_PLACEHOLDER_ID}"),
    ("get", f"/api/v1/reports/{_PLACEHOLDER_ID}/download"),
    ("get", "/api/v1/reports"),
]


@pytest.mark.parametrize("method,path", _UNAUTH_ROUTES, ids=[f"{m}:{p}" for m, p in _UNAUTH_ROUTES])
async def test_route_requires_authentication(auth_client, method, path):
    response = await getattr(auth_client, method)(path)
    assert response.status_code == 401


# --------------------------------------------------------------------------
# Authorization -- reports:read gates every read route
# --------------------------------------------------------------------------


async def test_dashboard_route_denied_without_permission(authenticated_client, harness):
    ac, user = authenticated_client
    harness.grant_role(user, org_id=uuid.uuid4(), role_code="sales_staff", permission_codes=[])
    response = await ac.get("/api/v1/dashboards/executive")
    assert response.status_code == 403


async def test_analytics_route_denied_without_permission(authenticated_client, harness):
    ac, user = authenticated_client
    harness.grant_role(user, org_id=uuid.uuid4(), role_code="sales_staff", permission_codes=[])
    response = await ac.get("/api/v1/analytics/kpi-summary")
    assert response.status_code == 403


async def test_report_catalog_denied_without_permission(authenticated_client, harness):
    ac, user = authenticated_client
    harness.grant_role(user, org_id=uuid.uuid4(), role_code="sales_staff", permission_codes=[])
    response = await ac.get("/api/v1/reports/catalog")
    assert response.status_code == 403


async def test_create_report_denied_with_only_reports_read(authenticated_client, harness):
    """`POST /reports` requires `reports:export`, not merely `reports:read`."""
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports", json={"report_type": "sales", "format": "csv"})
    assert response.status_code == 403


async def test_create_scheduled_report_denied_with_only_reports_read(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.post(
        "/api/v1/reports/scheduled",
        json={"name": "Weekly", "report_type": "sales", "format": "csv", "frequency": "weekly", "next_run_at": _future()},
    )
    assert response.status_code == 403


# --------------------------------------------------------------------------
# Dashboards -- 200 with reports:read, correct response shape
# --------------------------------------------------------------------------


async def test_executive_dashboard_returns_expected_shape(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get("/api/v1/dashboards/executive")
    assert response.status_code == 200
    body = response.json()
    assert set(["revenue_today", "revenue_mtd", "active_plant_count", "branches", "revenue_trend"]).issubset(body.keys())


async def test_nursery_dashboard_returns_expected_shape(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get("/api/v1/dashboards/nursery")
    assert response.status_code == 200
    assert response.json()["nursery_id"] == str(org_id)


async def test_branch_dashboard_404_for_unknown_branch(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get(f"/api/v1/dashboards/branch/{uuid.uuid4()}")
    assert response.status_code == 404


async def test_branch_dashboard_found_for_seeded_branch(authenticated_client, harness):
    from app.models.organization import Branch

    ac, user = authenticated_client
    org_id = uuid.uuid4()
    branch = Branch(
        id=uuid.uuid4(), nursery_id=org_id, name="Main", address_line1="1 St", city="Town",
        country="US", timezone="UTC", created_at=datetime.now(timezone.utc), updated_at=datetime.now(timezone.utc),
    )
    harness.branches.branches[branch.id] = branch
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get(f"/api/v1/dashboards/branch/{branch.id}")
    assert response.status_code == 200
    assert response.json()["branch_id"] == str(branch.id)


@pytest.mark.parametrize(
    "path", ["/api/v1/dashboards/plant", "/api/v1/dashboards/inventory", "/api/v1/dashboards/sales",
             "/api/v1/dashboards/customer", "/api/v1/dashboards/ai", "/api/v1/dashboards/financial"]
)
async def test_dashboard_routes_with_optional_branch_id_return_200(authenticated_client, harness, path):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get(path)
    assert response.status_code == 200


# --------------------------------------------------------------------------
# Analytics -- 200 with reports:read, date-range validation
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    "path", ["/api/v1/analytics/kpi-summary", "/api/v1/analytics/revenue-trend", "/api/v1/analytics/growth-trend",
             "/api/v1/analytics/inventory-trend", "/api/v1/analytics/plant-health-trend", "/api/v1/analytics/sales-forecast",
             "/api/v1/analytics/disease-trend", "/api/v1/analytics/customer-analytics",
             "/api/v1/analytics/employee-productivity", "/api/v1/analytics/branch-performance"]
)
async def test_analytics_routes_return_200(authenticated_client, harness, path):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get(path)
    assert response.status_code == 200


async def test_revenue_trend_rejects_inverted_date_range(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    now = datetime.now(timezone.utc)
    response = await ac.get(
        "/api/v1/analytics/revenue-trend",
        params={"date_from": now.isoformat(), "date_to": (now - timedelta(days=5)).isoformat()},
    )
    assert response.status_code == 422


async def test_growth_trend_accepts_species_id_filter(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get("/api/v1/analytics/growth-trend", params={"species_id": str(uuid.uuid4())})
    assert response.status_code == 200


# --------------------------------------------------------------------------
# Report catalog
# --------------------------------------------------------------------------


async def test_report_catalog_lists_every_report_type(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get("/api/v1/reports/catalog")
    assert response.status_code == 200
    body = response.json()
    assert len(body) == len(ReportType)
    assert all(entry["title"] and entry["description"] for entry in body)


# --------------------------------------------------------------------------
# Report generation / status / download / history
# --------------------------------------------------------------------------


async def test_create_report_returns_202_and_completes_synchronously(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)

    create_response = await ac.post("/api/v1/reports", json={"report_type": "branch", "format": "json"})
    assert create_response.status_code == 202
    body = create_response.json()
    assert body["status"] == "pending"
    assert body["nursery_id"] == str(org_id)
    assert body["requested_by_user_id"] == str(user.id)

    # `BackgroundTasks` execute within the same ASGI call in this test
    # transport (no real worker/queue involved) -- by the time the
    # response above was received, generation has already finished.
    status_response = await ac.get(f"/api/v1/reports/{body['id']}")
    assert status_response.status_code == 200
    status_body = status_response.json()
    assert status_body["status"] == "complete"
    assert status_body["download_url"] == f"/reports/{body['id']}/download"


async def test_create_report_never_trusts_client_supplied_nursery_id(authenticated_client, harness):
    """`ReportCreateRequest` has no `nursery_id` field at all -- the created report's `nursery_id` is always `tenant.org_id`, confirmed here even though nothing in the request body could try to override it."""
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports", json={"report_type": "branch", "format": "json"})
    assert response.json()["nursery_id"] == str(org_id)


async def test_create_report_rejects_invalid_report_type(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports", json={"report_type": "not_a_real_type", "format": "csv"})
    assert response.status_code == 422


async def test_create_report_rejects_invalid_format(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports", json={"report_type": "sales", "format": "not_a_real_format"})
    assert response.status_code == 422


async def test_create_report_rejects_inverted_filter_date_range(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    now = datetime.now(timezone.utc)
    response = await ac.post(
        "/api/v1/reports",
        json={
            "report_type": "sales", "format": "csv",
            "filters": {"date_from": now.isoformat(), "date_to": (now - timedelta(days=1)).isoformat()},
        },
    )
    assert response.status_code == 422


async def test_failed_report_status_and_error_never_crashes_the_route(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)

    create_response = await ac.post(
        "/api/v1/reports", json={"report_type": "plant", "format": "csv", "filters": {"status": "not-a-real-status"}}
    )
    assert create_response.status_code == 202
    report_id = create_response.json()["id"]

    status_response = await ac.get(f"/api/v1/reports/{report_id}")
    assert status_response.status_code == 200
    body = status_response.json()
    assert body["status"] == "failed"
    assert body["download_url"] is None


async def test_download_returns_404_when_report_not_yet_complete(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    # Directly persist a still-PENDING report -- bypasses the (synchronous
    # in this test transport) background task entirely, so it genuinely
    # never reaches COMPLETE.
    from app.models.reports import Report

    report = await harness.reports.add(
        Report(nursery_id=org_id, branch_id=None, report_type=ReportType.SALES, format=ReportFormat.CSV,
               filters={}, status=ReportStatus.PENDING, requested_by_user_id=user.id)
    )
    response = await ac.get(f"/api/v1/reports/{report.id}/download")
    assert response.status_code == 404


async def test_download_returns_404_for_unknown_report(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.get(f"/api/v1/reports/{uuid.uuid4()}/download")
    assert response.status_code == 404


async def test_get_report_status_404_for_unknown_report(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.get(f"/api/v1/reports/{uuid.uuid4()}")
    assert response.status_code == 404


# --------------------------------------------------------------------------
# Report history -- pagination, filtering, tenant/branch isolation
# --------------------------------------------------------------------------


async def test_list_reports_pagination(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    for _ in range(5):
        await ac.post("/api/v1/reports", json={"report_type": "branch", "format": "json"})

    page1 = await ac.get("/api/v1/reports", params={"page": 1, "page_size": 2})
    assert page1.status_code == 200
    body1 = page1.json()
    assert body1["meta"]["total_items"] == 5
    assert body1["meta"]["total_pages"] == 3
    assert len(body1["items"]) == 2

    page3 = await ac.get("/api/v1/reports", params={"page": 3, "page_size": 2})
    assert len(page3.json()["items"]) == 1


async def test_list_reports_filters_by_report_type(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    await ac.post("/api/v1/reports", json={"report_type": "branch", "format": "json"})
    await ac.post("/api/v1/reports", json={"report_type": "sales", "format": "json"})

    response = await ac.get("/api/v1/reports", params={"report_type": "sales"})
    body = response.json()
    assert body["meta"]["total_items"] == 1
    assert body["items"][0]["report_type"] == "sales"


async def test_list_reports_newest_first(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    first = await ac.post("/api/v1/reports", json={"report_type": "branch", "format": "json"})
    second = await ac.post("/api/v1/reports", json={"report_type": "sales", "format": "json"})

    response = await ac.get("/api/v1/reports")
    items = response.json()["items"]
    assert items[0]["id"] == second.json()["id"]
    assert items[1]["id"] == first.json()["id"]


async def test_list_reports_is_scoped_to_the_callers_own_org(authenticated_client, harness):
    ac, user = authenticated_client
    own_org = uuid.uuid4()
    foreign_org = uuid.uuid4()
    _grant_export(harness, user, org_id=own_org)
    from app.models.reports import Report

    await harness.reports.add(
        Report(nursery_id=foreign_org, branch_id=None, report_type=ReportType.SALES, format=ReportFormat.CSV,
               filters={}, status=ReportStatus.COMPLETE, requested_by_user_id=uuid.uuid4())
    )

    response = await ac.get("/api/v1/reports")
    assert response.json()["meta"]["total_items"] == 0


async def test_get_report_status_cross_tenant_is_403_not_404(authenticated_client, harness):
    """Per app/api/routes/reports.py's own docstring -- an existing report belonging to another org is a 403, not a 404 (only a genuinely nonexistent id is a 404)."""
    ac, user = authenticated_client
    own_org = uuid.uuid4()
    foreign_org = uuid.uuid4()
    _grant_read(harness, user, org_id=own_org)
    from app.models.reports import Report

    foreign_report = await harness.reports.add(
        Report(nursery_id=foreign_org, branch_id=None, report_type=ReportType.SALES, format=ReportFormat.CSV,
               filters={}, status=ReportStatus.COMPLETE, requested_by_user_id=uuid.uuid4())
    )

    response = await ac.get(f"/api/v1/reports/{foreign_report.id}")
    assert response.status_code == 403


async def test_get_report_status_cross_branch_is_403(authenticated_client, harness):
    """A report scoped to a specific branch the caller's role is not scoped to is a 403, even within the caller's own org."""
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    my_branch = uuid.uuid4()
    other_branch = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id, branch_ids=[my_branch])
    from app.models.reports import Report

    report = await harness.reports.add(
        Report(nursery_id=org_id, branch_id=other_branch, report_type=ReportType.SALES, format=ReportFormat.CSV,
               filters={}, status=ReportStatus.COMPLETE, requested_by_user_id=uuid.uuid4())
    )

    response = await ac.get(f"/api/v1/reports/{report.id}")
    assert response.status_code == 403


async def test_get_report_status_org_wide_report_accessible_to_branch_scoped_user(authenticated_client, harness):
    """`report.branch_id=None` (an org-wide report) skips the branch-membership check entirely -- see `_authorize_report`'s/`_authorize_branch_or_org`'s own docstring; a branch-scoped user with `reports:read` can still see it."""
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    my_branch = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id, branch_ids=[my_branch])
    from app.models.reports import Report

    report = await harness.reports.add(
        Report(nursery_id=org_id, branch_id=None, report_type=ReportType.SALES, format=ReportFormat.CSV,
               filters={}, status=ReportStatus.COMPLETE, requested_by_user_id=uuid.uuid4())
    )

    response = await ac.get(f"/api/v1/reports/{report.id}")
    assert response.status_code == 200


async def test_dashboard_route_branch_filter_denied_for_out_of_scope_branch(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    my_branch = uuid.uuid4()
    other_branch = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id, branch_ids=[my_branch])

    response = await ac.get("/api/v1/dashboards/plant", params={"branch_id": str(other_branch)})
    assert response.status_code == 403


async def test_dashboard_route_branch_filter_allowed_for_own_branch(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    my_branch = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id, branch_ids=[my_branch])

    response = await ac.get("/api/v1/dashboards/plant", params={"branch_id": str(my_branch)})
    assert response.status_code == 200


# --------------------------------------------------------------------------
# Export formats -- all four, end to end through generation + download
# --------------------------------------------------------------------------


@pytest.mark.parametrize("format_value", ["pdf", "excel", "csv", "json"])
async def test_generate_and_download_every_export_format(authenticated_client, harness, format_value):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)

    create_response = await ac.post("/api/v1/reports", json={"report_type": "branch", "format": format_value})
    assert create_response.status_code == 202
    report_id = create_response.json()["id"]

    status_response = await ac.get(f"/api/v1/reports/{report_id}")
    assert status_response.json()["status"] == "complete"

    download_response = await ac.get(f"/api/v1/reports/{report_id}/download")
    assert download_response.status_code == 200
    content = download_response.content

    if format_value == "pdf":
        assert content.startswith(b"%PDF")
    elif format_value == "excel":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content))
        assert workbook.active.max_row >= 1
    elif format_value == "csv":
        rows = list(csv.reader(io.StringIO(content.decode("utf-8"))))
        assert rows[0] == ["id", "name", "city", "region", "country", "status", "phone", "email"]
    else:  # json
        payload = json.loads(content)
        assert payload["row_count"] == 0


# --------------------------------------------------------------------------
# Scheduled reports -- CRUD, pause/resume, delete
# --------------------------------------------------------------------------


def _create_scheduled_body(**overrides) -> dict:
    base = {
        "name": "Weekly Sales", "report_type": "sales", "format": "csv",
        "frequency": "weekly", "next_run_at": _future(),
    }
    base.update(overrides)
    return base


async def test_create_scheduled_report_201(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body())
    assert response.status_code == 201
    body = response.json()
    assert body["is_active"] is True
    assert body["nursery_id"] == str(org_id)
    assert body["created_by_user_id"] == str(user.id)


async def test_create_scheduled_report_rejects_past_next_run_at(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    past = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body(next_run_at=past))
    assert response.status_code == 422


async def test_create_scheduled_report_rejects_blank_name(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body(name=""))
    assert response.status_code == 422


async def test_create_scheduled_report_rejects_invalid_frequency(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body(frequency="hourly"))
    assert response.status_code == 422


async def test_list_scheduled_reports_pagination(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    for i in range(3):
        await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body(name=f"Sched {i}"))

    response = await ac.get("/api/v1/reports/scheduled", params={"page": 1, "page_size": 2})
    body = response.json()
    assert body["meta"]["total_items"] == 3
    assert len(body["items"]) == 2


async def test_get_update_pause_resume_delete_scheduled_report_lifecycle(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    create_response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body())
    scheduled_id = create_response.json()["id"]

    get_response = await ac.get(f"/api/v1/reports/scheduled/{scheduled_id}")
    assert get_response.status_code == 200
    assert get_response.json()["name"] == "Weekly Sales"

    update_response = await ac.patch(f"/api/v1/reports/scheduled/{scheduled_id}", json={"name": "Renamed Weekly"})
    assert update_response.status_code == 200
    assert update_response.json()["name"] == "Renamed Weekly"

    pause_response = await ac.post(f"/api/v1/reports/scheduled/{scheduled_id}/pause")
    assert pause_response.status_code == 200
    assert pause_response.json()["is_active"] is False

    resume_response = await ac.post(f"/api/v1/reports/scheduled/{scheduled_id}/resume")
    assert resume_response.status_code == 200
    assert resume_response.json()["is_active"] is True

    delete_response = await ac.delete(f"/api/v1/reports/scheduled/{scheduled_id}")
    assert delete_response.status_code == 204

    after_delete = await ac.get(f"/api/v1/reports/scheduled/{scheduled_id}")
    assert after_delete.status_code == 404


async def test_update_scheduled_report_rejects_past_next_run_at(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    create_response = await ac.post("/api/v1/reports/scheduled", json=_create_scheduled_body())
    scheduled_id = create_response.json()["id"]

    past = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    response = await ac.patch(f"/api/v1/reports/scheduled/{scheduled_id}", json={"next_run_at": past})
    assert response.status_code == 422


async def test_scheduled_report_get_404_for_unknown_id(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.get(f"/api/v1/reports/scheduled/{uuid.uuid4()}")
    assert response.status_code == 404


async def test_scheduled_report_cross_tenant_is_403_not_404(authenticated_client, harness):
    ac, user = authenticated_client
    own_org = uuid.uuid4()
    foreign_org = uuid.uuid4()
    _grant_read(harness, user, org_id=own_org)
    foreign_scheduled = await harness.scheduled_report_service.create(
        nursery_id=foreign_org, branch_id=None, name="Foreign", report_type=ReportType.SALES, format=ReportFormat.CSV,
        filters={}, frequency=ReportScheduleFrequency.DAILY, created_by_user_id=uuid.uuid4(),
        next_run_at=datetime.now(timezone.utc) + timedelta(days=1),
    )

    response = await ac.get(f"/api/v1/reports/scheduled/{foreign_scheduled.id}")
    assert response.status_code == 403


async def test_scheduled_report_cross_branch_is_403(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    my_branch = uuid.uuid4()
    other_branch = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id, branch_ids=[my_branch])
    scheduled = await harness.scheduled_report_service.create(
        nursery_id=org_id, branch_id=other_branch, name="Other Branch", report_type=ReportType.SALES, format=ReportFormat.CSV,
        filters={}, frequency=ReportScheduleFrequency.DAILY, created_by_user_id=uuid.uuid4(),
        next_run_at=datetime.now(timezone.utc) + timedelta(days=1),
    )

    response = await ac.get(f"/api/v1/reports/scheduled/{scheduled.id}")
    assert response.status_code == 403


async def test_delete_scheduled_report_requires_reports_export_not_just_read(authenticated_client, harness):
    """`DELETE /reports/scheduled/{id}` is gated on `reports:export`, not merely `reports:read` -- a read-only caller must be denied even for a schedule that genuinely exists and belongs to their own org."""
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    scheduled = await harness.scheduled_report_service.create(
        nursery_id=org_id, branch_id=None, name="Weekly", report_type=ReportType.SALES, format=ReportFormat.CSV,
        filters={}, frequency=ReportScheduleFrequency.WEEKLY, created_by_user_id=uuid.uuid4(),
        next_run_at=datetime.now(timezone.utc) + timedelta(days=1),
    )
    _grant_read(harness, user, org_id=org_id)

    response = await ac.delete(f"/api/v1/reports/scheduled/{scheduled.id}")
    assert response.status_code == 403


# --------------------------------------------------------------------------
# Run-due -- idempotent execution, cross-tenant isolation, paused/deleted-never-execute
# --------------------------------------------------------------------------


async def test_run_due_executes_only_the_callers_own_org(authenticated_client, harness):
    ac, user = authenticated_client
    own_org = uuid.uuid4()
    foreign_org = uuid.uuid4()
    _grant_export(harness, user, org_id=own_org)

    own_scheduled = await harness.scheduled_report_service.create(
        nursery_id=own_org, branch_id=None, name="Mine", report_type=ReportType.BRANCH, format=ReportFormat.JSON,
        filters={}, frequency=ReportScheduleFrequency.DAILY, created_by_user_id=user.id,
        next_run_at=datetime.now(timezone.utc) + timedelta(days=1),
    )
    foreign_scheduled = await harness.scheduled_report_service.create(
        nursery_id=foreign_org, branch_id=None, name="Not Mine", report_type=ReportType.BRANCH, format=ReportFormat.JSON,
        filters={}, frequency=ReportScheduleFrequency.DAILY, created_by_user_id=uuid.uuid4(),
        next_run_at=datetime.now(timezone.utc) + timedelta(days=1),
    )
    # Force both due directly against the fake repo (bypasses the create-time "not in the past" validator).
    harness.scheduled_reports.scheduled[own_scheduled.id].next_run_at = datetime.now(timezone.utc) - timedelta(minutes=1)
    harness.scheduled_reports.scheduled[foreign_scheduled.id].next_run_at = datetime.now(timezone.utc) - timedelta(minutes=1)
    foreign_next_run_before = harness.scheduled_reports.scheduled[foreign_scheduled.id].next_run_at

    response = await ac.post("/api/v1/reports/scheduled/run-due")

    assert response.status_code == 200
    body = response.json()
    assert body["executed_count"] == 1
    assert body["results"][0]["scheduled_report_id"] == str(own_scheduled.id)
    # The foreign org's due schedule was never touched.
    assert harness.scheduled_reports.scheduled[foreign_scheduled.id].next_run_at == foreign_next_run_before


async def test_run_due_never_executes_a_paused_schedule(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    create_response = await ac.post(
        "/api/v1/reports/scheduled", json=_create_scheduled_body(report_type="branch", format="json")
    )
    scheduled_id = create_response.json()["id"]
    await ac.post(f"/api/v1/reports/scheduled/{scheduled_id}/pause")
    harness.scheduled_reports.scheduled[uuid.UUID(scheduled_id)].next_run_at = datetime.now(timezone.utc) - timedelta(minutes=1)

    response = await ac.post("/api/v1/reports/scheduled/run-due")

    assert response.json()["executed_count"] == 0


async def test_run_due_requires_reports_export_permission(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_read(harness, user, org_id=org_id)
    response = await ac.post("/api/v1/reports/scheduled/run-due")
    assert response.status_code == 403


async def test_run_due_is_idempotent_within_the_same_request(authenticated_client, harness):
    ac, user = authenticated_client
    org_id = uuid.uuid4()
    _grant_export(harness, user, org_id=org_id)
    create_response = await ac.post(
        "/api/v1/reports/scheduled", json=_create_scheduled_body(report_type="branch", format="json")
    )
    scheduled_id = uuid.UUID(create_response.json()["id"])
    harness.scheduled_reports.scheduled[scheduled_id].next_run_at = datetime.now(timezone.utc) - timedelta(minutes=1)

    first = await ac.post("/api/v1/reports/scheduled/run-due")
    second = await ac.post("/api/v1/reports/scheduled/run-due")

    assert first.json()["executed_count"] == 1
    assert second.json()["executed_count"] == 0  # already advanced into the future by the first call
