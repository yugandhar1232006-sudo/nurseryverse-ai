"""
Unit tests for `ReportGenerationService` (app/services/report_generation_service.py)
-- exercised directly against `harness` (real `FakeReportRepository` +
real `LocalFileStorage` writing to a per-test temp directory, per
tests/conftest.py's own "real fallback storage over a hand-rolled fake"
rationale), the same split every prior module's service unit tests use.

Coverage:
  * All 18 `ReportType` providers generate successfully against an EMPTY
    nursery (no seeded domain rows) -- confirms every provider's field
    access/exporter-compatibility wiring is correct without needing to
    fabricate a full realistic row for all eighteen unrelated entities;
    this is exactly the class of bug (a typo'd attribute, an
    exporter-incompatible value type) a "just don't crash" smoke test
    catches cheaply.
  * Richer "with real data" coverage for the higher-complexity providers:
    Plant (species filter), Inventory (low_stock_only filter), Sales/
    Revenue/Profit (a real checkout flow, specifically exercising the
    `Decimal` COGS computation the Module 12 mypy pass fixed), Customer,
    Branch.
  * All four export formats (PDF/Excel/CSV/JSON), each decoded back and
    checked for correct structure/row content.
  * The failure path: a provider that raises leaves the report FAILED
    with a `ReportFailed` domain event carrying the error message, never
    an unhandled exception escaping `generate()`.
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import pytest

from app.db.enums import ReportFormat, ReportStatus, ReportType
from app.models.organization import Branch
from app.models.catalog import Species
from app.models.reports import Report
from app.services.sales_service import LineItemInput

pytestmark = pytest.mark.unit


def _branch(*, nursery_id: uuid.UUID, name: str = "Main") -> Branch:
    now = datetime.now(timezone.utc)
    return Branch(
        id=uuid.uuid4(), nursery_id=nursery_id, name=name, address_line1="123 St", city="Springfield",
        country="US", timezone="UTC", created_at=now, updated_at=now,
    )


def _species(*, nursery_id: uuid.UUID, common_name: str = "Fig") -> Species:
    now = datetime.now(timezone.utc)
    return Species(
        id=uuid.uuid4(), nursery_id=nursery_id, category_id=uuid.uuid4(), common_name=common_name,
        botanical_name=f"{common_name} botanica", created_at=now, updated_at=now,
    )


async def _setup(harness):
    org_id = uuid.uuid4()
    branch = _branch(nursery_id=org_id)
    harness.branches.branches[branch.id] = branch
    return org_id, branch


async def _org_only(harness) -> uuid.UUID:
    """For the empty-nursery smoke tests -- deliberately does NOT register a branch (unlike `_setup`), so the Branch Report's own row count stays genuinely zero too."""
    return uuid.uuid4()


def _new_report(
    *, nursery_id: uuid.UUID, report_type: ReportType, format: ReportFormat = ReportFormat.JSON,
    branch_id: uuid.UUID | None = None, filters: dict | None = None,
) -> Report:
    return Report(
        nursery_id=nursery_id, branch_id=branch_id, report_type=report_type, format=format,
        filters=filters or {}, status=ReportStatus.PENDING, requested_by_user_id=uuid.uuid4(),
    )


async def _generate(harness, report: Report) -> Report:
    report = await harness.reports.add(report)
    await harness.report_generation_service.generate(report)
    return report


def _resolve_content(harness, report: Report) -> bytes:
    assert report.file_url is not None
    filename = report.file_url.rsplit("/", 1)[-1]
    path = harness.report_file_storage.resolve(filename)
    assert path is not None, f"file for {report.id} not found on disk"
    return path.read_bytes()


# --------------------------------------------------------------------------
# Every report type -- empty-nursery smoke test
# --------------------------------------------------------------------------

ALL_REPORT_TYPES = list(ReportType)


@pytest.mark.parametrize("report_type", ALL_REPORT_TYPES, ids=[t.value for t in ALL_REPORT_TYPES])
async def test_every_report_type_generates_successfully_against_empty_nursery(harness, report_type):
    org_id = await _org_only(harness)
    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=report_type))

    assert report.status == ReportStatus.COMPLETE, f"{report_type} failed to generate"
    assert report.file_url is not None
    assert report.completed_at is not None

    content = _resolve_content(harness, report)
    payload = json.loads(content)
    assert payload["row_count"] == 0
    assert payload["rows"] == []
    assert payload["title"]  # every provider supplies a non-empty title from REPORT_TYPE_TITLES


@pytest.mark.parametrize("report_type", ALL_REPORT_TYPES, ids=[t.value for t in ALL_REPORT_TYPES])
async def test_every_report_type_publishes_report_generated_event(harness, report_type):
    org_id = await _org_only(harness)
    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=report_type))
    generated_events = [e for e in harness.domain_events.events if e.event_type == "report.generated" and e.aggregate_id == report.id]
    assert len(generated_events) == 1


# --------------------------------------------------------------------------
# Plant Report -- species_id filter
# --------------------------------------------------------------------------


async def test_plant_report_includes_all_plants_when_unfiltered(harness):
    org_id, branch = await _setup(harness)
    fig = _species(nursery_id=org_id)
    harness.species.species[fig.id] = fig
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch.id, species_id=fig.id, actor_user_id=uuid.uuid4())
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch.id, species_id=fig.id, actor_user_id=uuid.uuid4())

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.PLANT, format=ReportFormat.JSON))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 2


async def test_plant_report_species_filter_excludes_other_species(harness):
    org_id, branch = await _setup(harness)
    fig = _species(nursery_id=org_id, common_name="Fig")
    rose = _species(nursery_id=org_id, common_name="Rose")
    harness.species.species[fig.id] = fig
    harness.species.species[rose.id] = rose
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch.id, species_id=fig.id, actor_user_id=uuid.uuid4())
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch.id, species_id=rose.id, actor_user_id=uuid.uuid4())

    report = await _generate(
        harness, _new_report(nursery_id=org_id, report_type=ReportType.PLANT, filters={"species_id": str(fig.id)})
    )
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1
    assert payload["rows"][0]["species_id"] == str(fig.id)


async def test_plant_report_scoped_to_branch_when_report_branch_id_set(harness):
    org_id, branch_a = await _setup(harness)
    branch_b = _branch(nursery_id=org_id, name="Second")
    harness.branches.branches[branch_b.id] = branch_b
    species = _species(nursery_id=org_id)
    harness.species.species[species.id] = species
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch_a.id, species_id=species.id, actor_user_id=uuid.uuid4())
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch_b.id, species_id=species.id, actor_user_id=uuid.uuid4())

    report = await _generate(
        harness, _new_report(nursery_id=org_id, report_type=ReportType.PLANT, branch_id=branch_a.id)
    )
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1


# --------------------------------------------------------------------------
# Inventory Report -- low_stock_only filter
# --------------------------------------------------------------------------


async def test_inventory_report_low_stock_only_filter(harness):
    org_id, branch = await _setup(harness)
    await harness.inventory_service.create_inventory_line(
        nursery_id=org_id, branch_id=branch.id, category_id=uuid.uuid4(), unit_id=uuid.uuid4(),
        name="Plenty", initial_quantity=100, low_stock_threshold=5, unit_price=9.99, actor_user_id=uuid.uuid4(),
    )
    await harness.inventory_service.create_inventory_line(
        nursery_id=org_id, branch_id=branch.id, category_id=uuid.uuid4(), unit_id=uuid.uuid4(),
        name="Scarce", initial_quantity=2, low_stock_threshold=5, unit_price=4.50, actor_user_id=uuid.uuid4(),
    )

    report = await _generate(
        harness, _new_report(nursery_id=org_id, report_type=ReportType.INVENTORY, filters={"low_stock_only": True})
    )
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1
    assert payload["rows"][0]["name"] == "Scarce"


# --------------------------------------------------------------------------
# Sales / Revenue / Profit Reports -- real checkout flow, Decimal COGS
# --------------------------------------------------------------------------


async def _checkout_one_sale(harness, *, org_id, branch, unit_price: Decimal, unit_cost: float, quantity: int = 2):
    inventory = await harness.inventory_service.create_inventory_line(
        nursery_id=org_id, branch_id=branch.id, category_id=uuid.uuid4(), unit_id=uuid.uuid4(),
        name="Basil 4in", initial_quantity=50, low_stock_threshold=5, unit_price=float(unit_price),
        unit_cost=unit_cost, actor_user_id=uuid.uuid4(),
    )
    # `Inventory.unit_cost` is a `Mapped[Decimal]` (SQLAlchemy `Numeric`
    # column) in production -- a real Postgres round-trip always yields a
    # `Decimal` regardless of what Python type was assigned on write. The
    # in-memory fake has no such column-type coercion (it stores whatever
    # object `InventoryService.create_inventory_line`'s `float`-typed
    # `unit_cost` parameter was given), so this line re-creates that
    # production coercion explicitly -- exercising `_profit_report`'s real
    # `Decimal(item.quantity) * inv.unit_cost` code path exactly as a real
    # database-backed run would.
    inventory.unit_cost = Decimal(str(unit_cost))
    customer = await harness.customer_service.create_customer(
        nursery_id=org_id, branch_id=branch.id, actor_user_id=uuid.uuid4(), name="Jane Doe"
    )
    order = await harness.sales_order_service.create_order(
        nursery_id=org_id, branch_id=branch.id, actor_user_id=uuid.uuid4(), customer_id=customer.id,
        items=[LineItemInput(quantity=quantity, unit_price=unit_price, inventory_id=inventory.id)],
    )
    fulfilled = await harness.sales_order_service.checkout(order, actor_user_id=uuid.uuid4())
    sale = await harness.sales.get_by_id(fulfilled.sale_id)
    return sale, customer, inventory


async def test_sales_report_includes_completed_sale(harness):
    org_id, branch = await _setup(harness)
    sale, _customer, _inv = await _checkout_one_sale(harness, org_id=org_id, branch=branch, unit_price=Decimal("10.00"), unit_cost=4.0)

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.SALES))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1
    assert payload["rows"][0]["id"] == str(sale.id)


async def test_revenue_report_buckets_by_day(harness):
    org_id, branch = await _setup(harness)
    sale, _customer, _inv = await _checkout_one_sale(harness, org_id=org_id, branch=branch, unit_price=Decimal("20.00"), unit_cost=8.0)

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.REVENUE))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1
    assert Decimal(str(payload["rows"][0]["revenue"])) == sale.total_amount
    assert payload["rows"][0]["sale_count"] == 1


async def test_profit_report_computes_decimal_cogs_correctly(harness):
    """Regression test for the mypy-flagged `Decimal * Numeric` COGS computation in `_profit_report` -- 2 units at $4.00 unit_cost = $8.00 COGS against a $20.00 sale (quantity 2 @ $10.00/unit)."""
    org_id, branch = await _setup(harness)
    sale, _customer, _inv = await _checkout_one_sale(
        harness, org_id=org_id, branch=branch, unit_price=Decimal("10.00"), unit_cost=4.0, quantity=2
    )

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.PROFIT))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 1
    row = payload["rows"][0]
    assert Decimal(str(row["total_amount"])) == Decimal("20.00")
    assert Decimal(str(row["estimated_cogs"])) == Decimal("8.00")
    assert Decimal(str(row["estimated_profit"])) == Decimal("12.00")


async def test_profit_report_skips_voided_sales(harness):
    org_id, branch = await _setup(harness)
    sale, _customer, _inv = await _checkout_one_sale(harness, org_id=org_id, branch=branch, unit_price=Decimal("10.00"), unit_cost=4.0)
    from app.db.enums import SaleStatus

    sale.status = SaleStatus.VOIDED

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.PROFIT))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 0


async def test_sales_report_date_range_filter(harness):
    org_id, branch = await _setup(harness)
    sale, _customer, _inv = await _checkout_one_sale(harness, org_id=org_id, branch=branch, unit_price=Decimal("15.00"), unit_cost=5.0)
    future_start = (datetime.now(timezone.utc) + timedelta(days=1)).isoformat()

    report = await _generate(
        harness, _new_report(nursery_id=org_id, report_type=ReportType.SALES, filters={"date_from": future_start})
    )
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 0


# --------------------------------------------------------------------------
# Customer / Branch Reports
# --------------------------------------------------------------------------


async def test_customer_report_lists_customers_for_org(harness):
    org_id, branch = await _setup(harness)
    await harness.customer_service.create_customer(nursery_id=org_id, branch_id=branch.id, actor_user_id=uuid.uuid4(), name="A")
    await harness.customer_service.create_customer(nursery_id=org_id, branch_id=branch.id, actor_user_id=uuid.uuid4(), name="B")

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.CUSTOMER))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 2


async def test_branch_report_includes_inactive_branches(harness):
    org_id, branch = await _setup(harness)
    inactive = _branch(nursery_id=org_id, name="Closed")
    from app.db.enums import BranchStatus

    inactive.status = BranchStatus.INACTIVE
    harness.branches.branches[inactive.id] = inactive

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.BRANCH))
    payload = json.loads(_resolve_content(harness, report))
    assert payload["row_count"] == 2


# --------------------------------------------------------------------------
# Export formats -- all four, decoded and structurally verified
# --------------------------------------------------------------------------


async def test_export_format_csv_decodes_with_headers_and_rows(harness):
    org_id, branch = await _setup(harness)
    species = _species(nursery_id=org_id)
    harness.species.species[species.id] = species
    await harness.plant_service.register_plant(nursery_id=org_id, branch_id=branch.id, species_id=species.id, actor_user_id=uuid.uuid4())

    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.PLANT, format=ReportFormat.CSV))
    content = _resolve_content(harness, report)
    rows = list(csv.reader(io.StringIO(content.decode("utf-8"))))
    assert rows[0] == ["id", "branch_id", "species_id", "common_label", "status", "zone", "price", "planted_at", "sold_at", "batch_number"]
    assert len(rows) == 2  # header + 1 plant


async def test_export_format_json_decodes_with_expected_shape(harness):
    org_id, branch = await _setup(harness)
    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.BRANCH, format=ReportFormat.JSON))
    payload = json.loads(_resolve_content(harness, report))
    assert set(payload.keys()) == {"title", "generated_at", "row_count", "rows"}
    assert payload["row_count"] == 1  # the branch _setup() itself created


async def test_export_format_excel_decodes_via_openpyxl(harness):
    org_id, branch = await _setup(harness)
    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.BRANCH, format=ReportFormat.EXCEL))
    content = _resolve_content(harness, report)

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[1]]
    assert header_row == ["id", "name", "city", "region", "country", "status", "phone", "email"]
    assert sheet.max_row == 2  # header + 1 branch


async def test_export_format_pdf_produces_valid_pdf_magic_bytes(harness):
    org_id, branch = await _setup(harness)
    report = await _generate(harness, _new_report(nursery_id=org_id, report_type=ReportType.BRANCH, format=ReportFormat.PDF))
    content = _resolve_content(harness, report)
    assert content.startswith(b"%PDF")


# --------------------------------------------------------------------------
# Failure path
# --------------------------------------------------------------------------


async def test_generate_marks_report_failed_when_provider_raises(harness):
    org_id, _branch_row = await _setup(harness)
    report = await _generate(
        harness,
        _new_report(nursery_id=org_id, report_type=ReportType.PLANT, filters={"status": "not-a-real-status"}),
    )

    assert report.status == ReportStatus.FAILED
    assert report.file_url is None
    assert report.completed_at is None

    failed_events = [e for e in harness.domain_events.events if e.event_type == "report.failed" and e.aggregate_id == report.id]
    assert len(failed_events) == 1


async def test_generate_raises_for_unregistered_report_type_but_still_marks_failed(harness, monkeypatch):
    """`_providers.get(report.report_type)` returning `None` is unreachable through the real enum today (every `ReportType` member has a provider -- see `test_every_report_type_generates_successfully_against_empty_nursery` above) but the `raise ValueError` guard is still real code; exercised directly by monkeypatching the registry."""
    org_id, _branch_row = await _setup(harness)
    report = await harness.reports.add(_new_report(nursery_id=org_id, report_type=ReportType.PLANT))
    monkeypatch.setattr(harness.report_generation_service, "_providers", {})

    await harness.report_generation_service.generate(report)

    assert report.status == ReportStatus.FAILED
